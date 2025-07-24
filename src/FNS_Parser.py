from openpyxl import load_workbook
from typing import Union
import pandas as pd
import requests
import time
import random


def safe_get_arg(data: dict, keys: list[str]) -> int | str:
    """Извлекает вложенное значение из словаря или возвращает 'undefined' при ошибке.

    Args:
        data: Словарь с данными.
        keys: Ключи для поиска.

    Returns:
        int | str:
            - Найденное значение (int для аргумента).
            - 'undefined', если ключ отсутствует.

    Примеры:
        safe_get_arg({"a": {"b": 52}}, ["a", "b"])
        52
        safe_get_arg({"year": {"days": 2023}}, ["year", "a"])
        undefined
    """
    try:
        data = data[keys[0]][keys[1]]
    except KeyError:
        return 'undefined'

    return int(data)


def safe_get_interest_expense(data: dict, keys: list[str]) -> int | str:
    """Извлекает вложенное значение из словаря или возвращает 0 при ошибке.

        Args:
            data: Словарь с данными.
            keys: Ключи для поиска.

        Returns:
            int:
                - Найденное значение (int для аргумента).
                - 0, если ключ отсутствует.

        Примеры:
            safe_get_interest_expense({"a": {"b": 52}}, ["a", "b"])
            52
            safe_get_interest_expense({"year": {"s": 2023}}, ["year", "a"])
            0
        """
    try:
        data = data[keys[0]][keys[1]]
    except KeyError:
        return 0

    return int(data)


def calculate_ebit(net_profit: int | str, income_tax: int | str, interest_expense: int | str) -> int | str:
    """
    Рассчитывает показатель EBIT (Earnings Before Interest and Taxes) по формуле:
    EBIT = Чистая прибыль + Налог на прибыль + Процентные расходы

    Args:
        net_profit: Чистая прибыль
        income_tax: Налог на прибыль
        interest_expense: Процентные расходы

    Returns:
        int: Рассчитанное значение EBIT
        'undefined': Если какие-то данные отсутствуют

    Примеры:
        calculate_ebit(1000, 200, 50)
        1250
        calculate_ebit(500_000, 100_000, 'undefined')
        undefined
    """
    if not all(isinstance(x, int) for x in [net_profit, income_tax, interest_expense]):
        return 'undefined'

    ebit = net_profit + income_tax + interest_expense
    return ebit


def safe_get_address_or_year(data: dict, key_list: list[str]) -> int | str:
    """Извлекает вложенное значение из словаря или возвращает 'undefined' при ошибке.

    В зависимости от длины key_list:
    - Если 2 ключа: возвращает значение по пути key1 → key2 (адрес).
    - Если 1 ключ: возвращает число по ключу, уменьшенное на 1 (год для которого считаем EBIT).

    Args:
        data: Словарь с данными.
        key_list: Список ключей для поиска (1 или 2 элемента).

    Returns:
        int | str:
            - Найденное значение (int для года, str для адреса).
            - 'undefined', если ключ отсутствует.

    Примеры:
        safe_get_address_or_year({"a": {"b": "адрес"}}, ["a", "b"])
        'адрес'
        safe_get_address_or_year({"year": "2023"}, ["year"])
        2022
    """
    if len(key_list) == 2:
        try:
            for key in key_list:
                data = data[key]
        except KeyError:
            return 'undefined'

        return data
    else:
        try:
            data = int(data[key_list[0]][:4]) - 1
        except KeyError:
            return 'undefined'

        return data


def append_to_excel(data: list[list[Union[int, str]]]) -> None:
    """Добавляет данные в файл Excel, начиная с текущей строки.

    Функция записывает переданные данные в файл 'ParsData.xlsx' в активный лист,
    начиная с ячейки A{row_number}. Каждый внутренний список записывается в отдельную строку,
    последовательно заполняя колонки от A до G. После записи автоматически увеличивает
    глобальный счетчик строк row_number.

    Args:
        data (list[list[Union[int, str]]]): Список строк для записи, где каждая строка
            должна содержать 7 элементов (для колонок A-G). Поддерживаются значения
            типа int и str
    """
    global row_number
    wb = load_workbook("ParsData.xlsx")
    ws = wb.active
    for row in data:
        ws[f'A{row_number}'] = row[0]
        ws[f'B{row_number}'] = row[1]
        ws[f'C{row_number}'] = row[2]
        ws[f'D{row_number}'] = row[3]
        ws[f'E{row_number}'] = row[4]
        ws[f'F{row_number}'] = row[5]
        ws[f'G{row_number}'] = row[6]
        row_number += 1
    wb.save("ParsData.xlsx")


row_number = 1 #Строка, с которой нужно начать заполнение файла
headers = {
    "User-Agent": "",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
    "Accept-Encoding": "gzip, deflate, br, zstd",
    "Connection": "keep-alive",
    "Cache-Control": "no-cache",
    "Pragma": "no-cache",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "same-origin",
    "Sec-Fetch-User": "?1",
    "Upgrade-Insecure-Requests": "1",
    "Referer": "https://bo.nalog.ru/",
    "DNT": "1",
}
df = pd.read_excel('Data_Lab1.xlsx')
session = requests.Session()
for i in df['ИНН']:
    lst = []
    res = session.get(f"https://bo.nalog.gov.ru/advanced-search/organizations/search?query={i}&page=0&size=20",
                      headers=headers).json()
    time.sleep(2)

    try:
        org_id = res["content"][0]["id"]
        response = session.get(f"https://bo.nalog.gov.ru/nbo/organizations/{org_id}/bfo/", timeout=10,
                               headers=headers).json()
    except (KeyError, IndexError):
        lst.append([int(i), 0, 0, 0, 0, 0, 0])
        append_to_excel(lst)
        continue

    for j in response:
        net_prof = safe_get_arg(j['typeCorrections'][0]['correction'], ['financialResult', 'current2400'])
        inc_tax = safe_get_arg(j['typeCorrections'][0]['correction'], ['financialResult', 'current2410'])
        int_exp = safe_get_interest_expense(j['typeCorrections'][0]['correction'], ['financialResult', 'current2330'])
        address = safe_get_address_or_year(j['typeCorrections'][0]['correction'], ['bfoOrganizationInfo', 'address'])
        year = safe_get_address_or_year(j['typeCorrections'][0]['correction'], ['datePresent'])
        lst.append([int(i), net_prof, inc_tax, int_exp, calculate_ebit(net_prof, inc_tax, int_exp), year, address])
        if year == 2020:
            net_prof = safe_get_arg(j['typeCorrections'][0]['correction'], ['financialResult', 'previous2400'])
            inc_tax = safe_get_arg(j['typeCorrections'][0]['correction'], ['financialResult', 'previous2410'])
            int_exp = safe_get_interest_expense(j['typeCorrections'][0]['correction'],['financialResult', 'previous2330'])
            lst.append([int(i), net_prof, inc_tax, int_exp, calculate_ebit(net_prof, inc_tax, int_exp), 2019, address])

    append_to_excel(lst)
    time.sleep(random.uniform(3, 7))
